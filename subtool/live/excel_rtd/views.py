import logging
import os
import pandas as pd
from string import ascii_uppercase
from django.core.urlresolvers import reverse
from django.shortcuts import redirect

from data.web.views import web_imports
from rivers.settings import BASE_DIR, SERVER_DIR
from statement.models import Statement, Position
from openpyxl import load_workbook
from openpyxl.styles import Font, Style
from subtool.live.excel_rtd.stat import ExcelRtdStatData

logger = logging.getLogger('views')
MAX_ROW = 500
EXCEL_FILE = os.path.join(SERVER_DIR, 'rtd.xlsx')
STOCK_NAMES = ('last', 'net_chg', 'volume', 'open', 'high', 'low')
STOCK_INDEX = ()
STOCK_CELL = {
    'last': 'A4', 'net_chg': 'C4', 'volume': 'I4', 'open': 'J4', 'high': 'K4', 'low': 'L4',
    'bid': 'D4', 'ask': 'F4'
}
OPTION_ORDER = (
    'option_code', 'bid', 'mark', 'ask', 'last',
    'delta', 'gamma', 'theta', 'vega',
    'impl_vol', 'prob_itm', 'prob_otm', 'prob_touch',
    'volume', 'open_int', 'intrinsic', 'extrinsic',
)
HOLD_OPTION = (
    'option_code', 'contract', 'strike', 'exp', 'dte', 'qty',
    'trade_price', 'trade_value', 'close_price', 'close_value',
    'profit_loss_$', 'profit_loss_%'
)
"""
Delta	Gamma	Theta	Vega	Impl Vol	Prob.ITM	Prob.OTM	Prob.Touch
Volume	Open.Int	Option Code	BID	BX	ASK	AX
"""
CALL_NAMES = (
    'delta', 'gamma', 'theta', 'vega', 'impl_vol', 'prob_itm', 'prob_otm',
    'prob_touch', 'volume', 'open_int', 'option_code', 'bid', 'bx', 'ask', 'ax'

)
"""
BID	BX	ASK	AX	Delta	Gamma	Theta	Vega	Impl Vol
Prob.ITM	Prob.OTM	Prob.Touch	Volume	Open.Int	Option Code
"""
PUT_NAMES = (
    'bid', 'bx', 'ask', 'ax', 'delta', 'gamma', 'theta', 'vega', 'impl_vol',
    'prob_itm', 'prob_otm', 'prob_touch', 'volume', 'open_int', 'option_code'

)
STAGE_NAMES = (
    'price', 'lt_stage', 'lt_amount',
    'e_stage', 'e_amount', 'gt_stage', 'gt_amount'
)
CONDITION_NAMES = (
    'name', 'condition', 'profit_loss'
)
CLOSE_STAT = (
    'std_value', 'in_std', 'out_std', 'above_std', 'below_std',
    'close_gain', 'close_loss', '60_day_move'
)


def set_header(sheet, key, value):
    sheet[key] = value.replace('_', ' ').capitalize()
    sheet[key].style = Style(font=Font(bold=True))


def set_value(sheet, key, value):
    sheet[key] = str(value)


def set_int(sheet, key, value):
    sheet[key] = int(value)


def set_float(sheet, key, value):
    sheet[key] = value
    sheet[key].number_format = '0.00'


def set_percent(sheet, key, value):
    sheet[key] = value
    sheet[key].number_format = '0.00%'


def set_date(sheet, key, value):
    sheet[key] = value
    sheet[key].number_format = 'yy-mm-dd'


def set_time(sheet, key, value):
    sheet[key] = value
    sheet[key].number_format = 'h:m:s'


def make_close0_open1_expr(qcut_range, close_open):
    temp = [float(n) for n in qcut_range[1:-1].split(', ')]
    return temp[0], temp[1], '=AND(%s<=%s,%s<=%s)' % (
        temp[0], close_open, close_open, temp[1]
    )


def make_expr(cond, price):
    temp = cond.replace('x', 'value').replace('y', 'value')
    result = '%s' % temp.format(value=price).replace(' ', '').replace('==', '=')
    items = result.split('<')
    if len(items) == 3:
        result = '=AND(%s<%s, %s<%s)' % (items[0], items[1], items[1], items[2])
    else:
        result = '=%s' % result

    return result


class StockCell(object):
    def __init__(self, symbol):
        self.last = '%s!%s' % (symbol, STOCK_CELL['last'])
        self.net_chg = '%s!%s' % (symbol, STOCK_CELL['net_chg'])
        self.volume = '%s!%s' % (symbol, STOCK_CELL['volume'])
        self.open = '%s!%s' % (symbol, STOCK_CELL['open'])
        self.high = '%s!%s' % (symbol, STOCK_CELL['high'])
        self.low = '%s!%s' % (symbol, STOCK_CELL['low'])
        self.bid = '%s!%s' % (symbol, STOCK_CELL['bid'])
        self.ask = '%s!%s' % (symbol, STOCK_CELL['ask'])


def excel_rtd_create(request):
    """
    1. get all positions
    2. create a sheet in rtd.xlxs
    3. for every symbol
    3a. put symbol data into it
    3b. get data from other sheet
    3c. make calculation cell
    4. save and auto open file
    :param request: request
    :return: render
    """
    logger.info('create analysis sheet for excel rtd')
    wb = load_workbook(filename=EXCEL_FILE)
    sheets = wb.get_sheet_names()

    # get statement/holding data from db
    # symbols = ['BA', 'HRB']
    symbols = [n for n in wb.get_sheet_names() if '_' not in n]

    # import web data
    web_imports(symbols)

    # calc stat data
    excel_stat = ExcelRtdStatData(symbols)
    excel_stat.get_data()
    stats = {}
    for symbol in symbols:
        logger.info('create sheet for symbol: %s' % symbol.upper())
        # if symbol not in ('AIG', 'BP'):
        #    print symbol
        #    continue

        stats[symbol] = excel_stat.get_symbol_stat(symbol)

    # set excel cell
    for symbol in symbols:
        # if symbol not in ('AIG', 'BP'):
        #     continue

        sheet_name = '_%s' % symbol.upper()
        if sheet_name in sheets:
            ws0 = wb[sheet_name]
            wb.remove_sheet(ws0)
        ws0 = wb.create_sheet(0)
        ws0.title = sheet_name

        logger.info('Create for symbol: %s' % symbol)

        # all stock cell into symbol worksheet
        stock1 = StockCell(symbol)

        # set yesterday close
        close0 = stats[symbol].close  # yesterday close
        move1 = analysis_set_movement(ws0, close0, stock1)
        trade = analysis_set_pl_open(ws0, stock1, move1)
        set_volume_hl(ws0, stats[symbol], stock1)
        set_open_to_close(ws0, stats[symbol], stock1, move1, trade)

        set_close_to_open(ws0, stats[symbol], stock1, move1, trade)

        set_header(ws0, 'I4', 'Note: use remain volume for coming possible move trend or reverse')
        set_header(ws0, 'I5', 'Note: use high low for estimate hl for now')

        set_header(ws0, 'I6', 'Note: price move between part is rare, mostly stay or go nearest')
        set_header(ws0, 'I7', 'Note: open to close, need for determine price action')
        set_header(ws0, 'I8', 'Note: use for better enter and exit price')
        set_header(ws0, 'I9', 'Note: can also be use as day trade indicator')

        set_decision(ws0)
        set_extra_decision(ws0)
        set_action(ws0)

        set_std_move(ws0, stats[symbol], move1)
        set_consec_move(ws0, stats[symbol])
        set_day60_pl(ws0, trade)

        set_history_price(ws0, stats[symbol], trade)

    wb._active_sheet_index = 1
    wb.save(EXCEL_FILE)

    return redirect(reverse('admin:app_list', args=('statement',)))


def set_history_price(ws, stat, trade):
    """
    Set history price with trade data
    :param ws: worksheet
    :param stat: StatDAta
    :param trade: TradeCell
    """
    row = 85
    # history price
    set_header(ws, 'A%d' % row, 'date')
    set_header(ws, 'B%d' % row, 'close')
    set_header(ws, 'C%d' % row, 'volume')
    set_header(ws, 'D%d' % row, 'close_chg')
    set_header(ws, 'E%d' % row, 'pct_chg')
    set_header(ws, 'F%d' % row, 'consec')
    set_header(ws, 'G%d' % row, 'out_std')
    set_header(ws, 'H%d' % row, 'above_std')
    set_header(ws, 'I%d' % row, 'below_std')
    set_header(ws, 'J%d' % row, 'trade_pl')

    row += 1
    for history in stat.std_close['history']:
        set_value(ws, 'A%d' % row, history['date'].strftime('%Y-%m-%d'))
        set_float(ws, 'B%d' % row, history['close'])
        set_int(ws, 'C%d' % row, history['volume'])
        set_float(ws, 'D%d' % row, history['close_chg'])
        set_percent(ws, 'E%d' % row, history['pct_chg'])
        set_int(ws, 'F%d' % row, history['consec'])
        set_value(ws, 'G%d' % row, '=IF(%s,TRUE,"")' % history['out_std'])
        set_value(ws, 'H%d' % row, '=IF(%s,TRUE,"")' % history['above_std'])
        set_value(ws, 'I%d' % row, '=IF(%s,TRUE,"")' % history['below_std'])
        set_percent(ws, 'J%d' % row, '=(B%d/%s)-1' % (row, trade.trade_price))

        row += 1


def set_day60_pl(ws, trade):
    """
    Set 60 days trade p/l
    :param ws: worksheet
    :param trade: TradeCell
    """
    row = 75
    history_pl = 'J%d:J%d' % (86, 145)

    set_header(ws, 'H%d' % row, 'item')
    set_header(ws, 'I%d' % row, 'stat')
    set_header(ws, 'J%d' % row, 'to open p/l')
    set_header(ws, 'K%d' % row, 'to last p/l')

    row += 1
    set_value(ws, 'H%d' % row, 'count P/L %')
    set_value(ws, 'I%d' % row, '=COUNT(%s)' % history_pl)

    row += 1
    set_value(ws, 'H%d' % row, 'profit %')
    set_percent(ws, 'I%d' % row, '=COUNTIF(%s,">0")/COUNT(%s)' % (history_pl, history_pl))

    row += 1
    set_value(ws, 'H%d' % row, 'loss %')
    set_percent(ws, 'I%d' % row, '=COUNTIF(%s,"<0")/COUNT(%s)' % (history_pl, history_pl))

    row += 1
    set_value(ws, 'H%d' % row, 'median P/L %')
    set_percent(ws, 'I%d' % row, '=MEDIAN(%s)' % history_pl)
    set_float(ws, 'J%d' % row, '=I%d*%s' % (row, trade.margin))
    set_float(ws, 'K%d' % row, '=J%d-%s' % (row, trade.pl_last))

    row += 1
    set_value(ws, 'H%d' % row, 'max %')
    set_percent(ws, 'I%d' % row, '=MAX(%s)' % history_pl)
    set_float(ws, 'J%d' % row, '=I%d*%s' % (row, trade.margin))
    set_float(ws, 'K%d' % row, '=J%d-%s' % (row, trade.pl_last))

    row += 1
    set_value(ws, 'H%d' % row, 'min %')
    set_percent(ws, 'I%d' % row, '=MIN(%s)' % history_pl)
    set_float(ws, 'J%d' % row, '=I%d*%s' % (row, trade.margin))
    set_float(ws, 'K%d' % row, '=J%d-%s' % (row, trade.pl_last))


def set_consec_move(ws, stat):
    """
    Set consec move
    :param ws: worksheet
    :param stat: StatData
    """
    # consec in std
    row = 75

    set_header(ws, 'E%d' % row, 'consec')
    set_header(ws, 'F%d' % row, 'count')
    row += 1
    for index, count in stat.std_close['consec']:
        set_value(ws, 'E%d' % row, '%d days in std' % index)
        set_int(ws, 'F%d' % row, count)
        row += 1


def set_std_move(ws, stat, move1):
    """
    Set std move
    :param ws: worksheet
    :param stat: StatData
    :param move1: MoveData
    """
    row = 75
    # stat predict close stat
    set_header(ws, 'A%d' % (row - 1), 'Statistics for pass 60 days')
    std_value = 'B%d' % (row + 1)
    set_header(ws, 'A%d' % row, 'item')
    set_header(ws, 'B%d' % row, 'stat')
    set_header(ws, 'C%d' % row, 'bool')
    row += 1
    for i, name in enumerate(CLOSE_STAT):
        set_header(ws, 'A%d' % (row + i), name)

        proc = set_int
        if name in ('std_value', '60_day_move'):
            proc = set_percent

        proc(
            ws, 'B%d' % (row + i), stat.std_close['std'][name]
        )

    # set bool
    expr = '=%s-ABS(%s)' % (std_value, move1.netChgPercent)  # std diff
    set_percent(ws, 'C%d' % row, expr)

    expr = '=IF(AND(-%s<=%s,%s<=%s),TRUE,"")' % (
        std_value, move1.netChgPercent, move1.netChgPercent, std_value
    )  # in std
    set_value(ws, 'C%d' % (row + 1), expr)
    expr = '=IF(OR(-%s>%s,%s>%s),TRUE,"")' % (
        std_value, move1.netChgPercent, move1.netChgPercent, std_value
    )  # out std
    set_value(ws, 'C%d' % (row + 2), expr)

    expr = '=IF(%s>%s,TRUE,"")' % (move1.netChgPercent, std_value)  # above std
    set_value(ws, 'C%d' % (row + 3), expr)
    expr = '=IF(%s<-%s,TRUE,"")' % (move1.netChgPercent, std_value)  # below std
    set_value(ws, 'C%d' % (row + 4), expr)

    expr = '=IF(%s>0,TRUE,"")' % move1.netChgPercent  # close gain
    set_value(ws, 'C%d' % (row + 5), expr)
    expr = '=IF(%s<0,TRUE,"")' % move1.netChgPercent  # close loss
    set_value(ws, 'C%d' % (row + 6), expr)


def set_decision(ws):
    """
    Set decision field
    :param ws: worksheet
    """
    row = 50
    set_header(ws, 'A%s' % (row - 1), 'Intraday decision maker')
    set_header(ws, 'A%s' % row, 'Item')
    set_header(ws, 'C%s' % row, 'Good')
    set_header(ws, 'D%s' % row, 'Action')

    set_header(ws, 'A%s' % (row + 1), 'Open P/L profit?')
    set_header(ws, 'A%s' % (row + 2), 'Open P/L more profit?')
    set_header(ws, 'A%s' % (row + 3), '50% time remain to run?')

    set_header(ws, 'A%s' % (row + 4), 'Remain volume profit?')
    set_header(ws, 'A%s' % (row + 5), 'H/L run all, safe now?')
    set_header(ws, 'A%s' % (row + 6), 'O2C remain profit?')
    set_header(ws, 'A%s' % (row + 7), 'O2C 30 minutes good?')
    set_header(ws, 'A%s' % (row + 8), 'O2C expect p/l good?')
    set_header(ws, 'A%s' % (row + 9), 'O2C move diff part?')
    set_header(ws, 'A%s' % (row + 10), 'C2O in profit group?')
    set_header(ws, 'A%s' % (row + 11), 'C2O will be in profit part?')
    set_header(ws, 'A%s' % (row + 12), 'C2O better hold?')
    set_header(ws, 'A%s' % (row + 13), 'Expect profit in std?')
    set_header(ws, 'A%s' % (row + 14), 'Expect profit out std?')

    for r in range(row + 1, row + 15):
        set_int(ws, 'C%d' % r, 0)
        set_value(ws, 'D%d' % r, '=IF(C%d, "Enter/Hold", "Exit")' % r)

    set_value(ws, 'C%d' % (row + 15), '=SUM(C%s:C%s)' % (row + 1, row + 14))
    set_percent(ws, 'D%d' % (row + 15), '=C%d/COUNT(C%s:C%s)' % (row + 15, row + 1, row + 14))


def set_extra_decision(ws):
    """
    Set extra decision field
    :param ws: worksheet
    """
    row = 50
    set_header(ws, 'F%d' % row, 'Item')
    set_header(ws, 'H%d' % row, 'Good')
    set_header(ws, 'I%d' % row, 'Action')

    set_header(ws, 'F%s' % (row + 1), 'Most trend to profit?')
    set_header(ws, 'F%s' % (row + 2), 'Active profit larger volume?')
    set_header(ws, 'F%s' % (row + 3), 'Level 2 moving to profit?')

    set_header(ws, 'F%s' % (row + 4), 'Option timesale to profit?')
    set_header(ws, 'F%s' % (row + 5), 'Timesale profit tomorrow?')

    set_header(ws, 'F%s' % (row + 6), 'Ticker 1 minute to profit?')
    set_header(ws, 'F%s' % (row + 7), 'Ticker 10 minute to profit?')
    set_header(ws, 'F%s' % (row + 8), 'Ticker 30 minute to profit?')

    set_header(ws, 'F%s' % (row + 9), '60days in profit?')
    set_header(ws, 'F%s' % (row + 10), '60days more profit?')
    set_header(ws, 'F%s' % (row + 11), '60days high chance?')

    for r in range(row + 1, row + 12):
        set_int(ws, 'H%d' % r, 0)
        set_value(ws, 'I%d' % r, '=IF(H%d, "Enter/Hold", "Exit")' % r)

    set_value(ws, 'H%d' % (row + 12), '=SUM(H%s:H%s)' % (row + 1, row + 11))
    set_percent(ws, 'I%d' % (row + 12), '=H%d/COUNT(H%s:H%s)' % (row + 12, row + 1, row + 11))


def set_action(ws):
    """
    Set action for this symbol except enter
    :param ws: worksheet
    """
    row = 67

    set_header(ws, 'A%d' % row, 'Trade action for intraday')

    actions = [
        (['A', 'C', 'D'], 'sell 50%', [
            'today loss- money?',
            'loss--- on upcoming week?',
            'loss--- no fd/news change?'
        ]),
        (['F', 'H', 'I'], 'exit 100%', [
            'today loss---?',
            'loss--- upcoming week?',
            'loss- fd/news change?',
        ]),
        (['K', 'M', 'N'], 'Buy more 50%', [
            'today profit++?',
            'profit++ upcoming week?',
            'profit+ & fd/news change?',
            'is long position?',
        ])
    ]

    # 'Close -100%',
    # 'Enter 100%',
    # 'Buy more +50%',
    row += 1
    for (c0, c1, c2), name, items in actions:
        set_header(ws, '%s%d' % (c0, row), name)
        length = len(items)
        for r, item in zip(range(row + 1, row + length + 1), items):
            set_header(ws, '%s%d' % (c0, r), item)
            set_int(ws, '%s%d' % (c1, r), 0)
            set_value(ws, '%s%d' % (c2, r), '=IF(%s%d, "Bad", "Good")' % (c1, r))

        bot_row = row + length + 1
        set_value(ws, '%s%d' % (c1, bot_row), '=SUM(%s%d:%s%d)' % (c1, row + 1, c1, bot_row - 1))
        set_percent(ws, '%s%d' % (c2, bot_row), '=%s%d/COUNT(%s%d:%s%d)' % (
            c1, bot_row, c1, row + 1, c1, bot_row - 1
        ))


def set_close_to_open(ws0, stat, stock1, move1, trade):
    """
    Set close to open statistic
    :param ws0: worksheet
    :param stat: StatData
    :param stock1: StockCell
    :param move1: MoveCell
    :param trade: TradeCell
    """
    row = 30
    set_header(ws0, 'A%s' % row, 'market move estimate after market')
    set_header(ws0, 'A%s' % (row + 1), 'Open to close move')
    set_header(ws0, 'D%s' % (row + 1), 'Close to open expect')
    set_header(ws0, 'A%s' % (row + 2), 'Min')
    set_header(ws0, 'B%s' % (row + 2), 'Max')
    set_header(ws0, 'C%s' % (row + 2), 'Bool')

    for i in range(3, 12, 2):
        set_header(ws0, '%s%s' % (ascii_uppercase[i], row + 2), 'Part %d' % (i - (i / 2) - 1))

    row += 3
    for stat in stat.co_stat:
        start0, stop0 = stat.group
        set_percent(ws0, 'A%d' % row, start0)
        set_percent(ws0, 'B%d' % row, stop0)

        expr0 = '=IF(AND(%s<=%s,%s<=%s), TRUE, "")' % (
            'A%d' % row, move1.o2lPercent, move1.o2lPercent, 'B%d' % row
        )
        expr_cell = 'C%d' % row
        set_value(ws0, expr_cell, expr0)

        for pos, (start1, stop1) in zip(range(3, 12, 2), stat.parts):
            percent0 = '%s%d' % (ascii_uppercase[pos], row)
            percent1 = '%s%d' % (ascii_uppercase[pos + 1], row)
            set_percent(ws0, percent0, start1)
            set_percent(ws0, percent1, stop1)

            price0 = '%s%d' % (ascii_uppercase[pos], row + 1)
            price1 = '%s%d' % (ascii_uppercase[pos + 1], row + 1)
            set_float(ws0, price0, '=IFERROR(IF(%s, (1+%s)*%s, ""), "")' % (
                expr_cell, percent0, stock1.last
            ))
            set_float(ws0, price1, '=IFERROR(IF(%s, (1+%s)*%s, ""), "")' % (
                expr_cell, percent1, stock1.last
            ))

            pl0 = '%s%d' % (ascii_uppercase[pos], row + 2)
            pl1 = '%s%d' % (ascii_uppercase[pos + 1], row + 2)
            set_float(ws0, pl0, '=IFERROR((%s-%s)*%s, "")' % (
                price0, trade.trade_price, trade.share
            ))
            set_float(ws0, pl1, '=IFERROR((%s-%s)*%s, "")' % (
                price1, trade.trade_price, trade.share
            ))

        row += 3


def set_open_to_close(ws0, stat, stock1, move1, trade):
    """
    Set open to close estimate statistics
    :param ws0: worksheet
    :param stat: StatData
    :param stock1: StockCell
    :param move1: MoveCell
    :param trade: TradeCell
    """
    # set open close header
    row = 12
    set_header(ws0, 'A%s' % row, 'before market estimate market move')
    set_header(ws0, 'A%s' % (row + 1), 'Close to open move')
    set_header(ws0, 'D%s' % (row + 1), 'Open to close expect')
    set_header(ws0, 'A%s' % (row + 2), 'Min')
    set_header(ws0, 'B%s' % (row + 2), 'Max')
    set_header(ws0, 'C%s' % (row + 2), 'Bool')

    for i in range(3, 8):
        set_header(ws0, '%s%d' % (ascii_uppercase[i], row + 2), 'Part %d' % (i - 2))

    set_header(ws0, 'I%d' % (row + 2), 'Using')
    set_header(ws0, 'J%d' % (row + 2), 'Price Range')
    set_header(ws0, 'K%d' % (row + 2), 'Remain %')
    set_header(ws0, 'L%d' % (row + 2), '% 30 Mins')
    set_header(ws0, 'M%d' % (row + 2), 'Expect P/L')

    # set value
    row = 15
    for stat in stat.oc_stat:
        # set min max header & value
        start0, stop0 = stat['close_open']
        set_percent(ws0, 'A%d' % row, start0)
        set_percent(ws0, 'B%d' % row, stop0)
        expr0 = '=IF(AND(%s<=%s,%s<=%s), TRUE, "")' % (
            'A%d' % row, move1.c2oPercent, move1.c2oPercent, 'B%d' % row
        )

        set_value(ws0, 'C%d' % row, expr0)

        for i, (start1, stop1) in enumerate(stat['open_close'], start=3):
            expr1 = '=IF(AND(%s<=%s,%s<=%s,C%d=TRUE), TRUE, "")' % (
                start1, move1.o2lPercent, move1.o2lPercent, stop1, row
            )

            set_percent(ws0, '%s%d' % (ascii_uppercase[i], row), start1)
            set_percent(ws0, '%s%d' % (ascii_uppercase[i], row + 1), stop1)
            set_value(ws0, '%s%d' % (ascii_uppercase[i], row + 2), expr1)

        row1 = row + 1
        row2 = row + 2

        search_index = '=IFERROR(INDEX(D%s:H%d, MATCH(TRUE, D%d:H%d)), "")'
        set_percent(ws0, 'I%d' % row, search_index % (row, row, row2, row2))
        set_percent(ws0, 'I%d' % row1, search_index % (row1, row1, row2, row2))

        price_range = '=IFERROR((1+I%d)*%s, "")'
        set_float(ws0, 'J%d' % row, price_range % (row, stock1.open))
        set_float(ws0, 'J%d' % row1, price_range % (row1, stock1.open))

        remain_pct = '=IFERROR(I%d-%s, "")'
        set_percent(ws0, 'K%d' % row, remain_pct % (row, move1.o2lPercent))
        set_percent(ws0, 'K%d' % row1, remain_pct % (row1, move1.o2lPercent))

        remain_value = '=IFERROR(K%d/%s*60*30, "")'
        set_percent(ws0, 'L%d' % row, remain_value % (row, move1.untilTime))
        set_percent(ws0, 'L%d' % row1, remain_value % (row1, move1.untilTime))

        expect_pl = '=IFERROR((J%d-%s)*%s, "")'
        set_float(ws0, 'M%d' % row, expect_pl % (row, trade.trade_price, trade.share))
        set_float(ws0, 'M%d' % row1, expect_pl % (row1, trade.trade_price, trade.share))

        row += 3


class MoveCell(object):
    def __init__(self):
        self.close0 = 'A2'
        self.c2oValue = 'B2'
        self.c2oPercent = 'C2'
        self.netChgPercent = 'E2'  # close to last
        self.o2lValue = 'F2'
        self.o2lPercent = 'G2'
        self.untilTime = 'J2'


def analysis_set_movement(ws0, close0, stock1):
    """
    Set stock price intraday movement
    :param ws0: worksheet
    :param close0: float
    :param stock1: StockCell
    :return: MoveCell
    """
    move1 = MoveCell()

    set_header(ws0, 'A1', 'close0')
    set_float(ws0, 'A2', close0)

    # set today move
    set_header(ws0, 'B1', 'close to open $')  # close to open $
    set_float(ws0, 'B2', '=%s-A2' % stock1.open)
    set_header(ws0, 'C1', 'close to open %')  # close to open %
    set_percent(ws0, 'C2', '=B2/A2')
    set_header(ws0, 'D1', 'net chg')  # net chg %
    set_float(ws0, 'D2', '=%s' % stock1.net_chg)
    set_header(ws0, 'E1', 'net chg %')  # net chg %
    set_percent(ws0, 'E2', '=%s/A2' % stock1.net_chg)
    set_header(ws0, 'F1', 'open to last $')  # close to last $
    set_float(ws0, 'F2', '=%s-%s' % (stock1.last, stock1.open))
    set_header(ws0, 'G1', 'open to last %')  # close to last %
    set_percent(ws0, 'G2', '=F2/%s' % stock1.open)
    set_header(ws0, 'H1', 'Now')  # time
    set_time(ws0, 'H2', '=NOW()')
    set_header(ws0, 'I1', 'Until')  # time
    set_time(ws0, 'I2', '=TODAY()+TIME(16,0,0)-NOW()')
    set_header(ws0, 'J1', 'In Secs')  # time
    set_value(ws0, 'J2', '=HOUR(I2)*3600+MINUTE(I2)*60+SECOND(I2)')

    return move1


class TradeCell(object):
    def __init__(self):
        self.trade_price = 'A5'
        self.share = 'B5'
        self.margin = 'C5'
        self.pl_last = 'F5'


def analysis_set_pl_open(ws0, stock1, move1):
    """
    Set Trade position, profit loss Cell
    :param ws0: worksheet
    :param stock1: StockCell
    :param move1: MoveCell
    :return: TradeCell
    """
    trade = TradeCell()

    set_header(ws0, 'A4', 'trade price')
    # set_float(ws0, 'A5', '=%s' % move1.close0)
    set_float(ws0, 'A5', '=%s' % 0)
    set_header(ws0, 'B4', 'shares')
    set_int(ws0, 'B5', 100)
    set_header(ws0, 'C4', 'margin')
    set_float(ws0, 'C5', '=A5*B5')
    set_header(ws0, 'D4', 'pl close0 $')
    set_float(ws0, 'D5', '=(%s-A5)*B5' % move1.close0)
    set_header(ws0, 'E4', 'pl close0 %')
    set_percent(ws0, 'E5', '=%s/A5-1' % move1.close0)
    set_header(ws0, 'F4', 'pl last $')
    set_float(ws0, 'F5', '=(%s-A5)*B5' % stock1.last)
    set_header(ws0, 'G4', 'pl last %')
    set_percent(ws0, 'G5', '=%s/A5-1' % stock1.last)

    return trade


def set_volume_hl(ws0, stat, stock1):
    """
    Set high low volume width
    :param ws0: worksheet
    :param stat: StatData
    :param stock1: StockCell
    """
    # set excel stat
    set_header(ws0, 'A7', 'days')
    set_header(ws0, 'B7', 'mean volume')
    set_header(ws0, 'C7', 'volume left')
    set_header(ws0, 'D7', 'volume left %')
    set_header(ws0, 'E7', 'mean hl')
    set_header(ws0, 'F7', 'current hl')
    set_header(ws0, 'G7', 'remain hl')

    # mean day 5
    set_int(ws0, 'A8', 5)
    set_int(ws0, 'B8', stat.mean_vol5)
    set_value(ws0, 'C8', '=B8-%s' % stock1.volume)
    set_percent(ws0, 'D8', '=C8/B8')
    set_percent(ws0, 'E8', stat.mean_hl5)
    set_percent(ws0, 'F8', '=(%s-%s)/%s' % (stock1.high, stock1.low, stock1.open))
    set_percent(ws0, 'G8', '=E8-F8')

    # mean day 20
    set_int(ws0, 'A9', 20)
    set_int(ws0, 'B9', stat.mean_vol20)
    set_value(ws0, 'C9', '=B9-%s' % stock1.volume)
    set_percent(ws0, 'D9', '=C9/B9')
    set_percent(ws0, 'E9', stat.mean_hl20)
    set_percent(ws0, 'F9', '=(%s-%s)/%s' % (stock1.high, stock1.low, stock1.open))
    set_percent(ws0, 'G9', '=E9-F9')


# no option yet, should add later
